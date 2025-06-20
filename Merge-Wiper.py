import os
import sys
import logging
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from colorama import init, Fore, Style
# Importa módulos para el navegador de archivos (interfaz gráfica)
import tkinter as tk
from tkinter import filedialog
# Importa módulos para reportes detallados (medición de recursos)
import time
import psutil

# Inicializa colorama para colorear la salida en consola automáticamente
init(autoreset=True)

# Configuración del sistema de logging para mostrar mensajes de depuración, información y errores
LOG_LEVEL = logging.DEBUG  # Puedes cambiar a logging.INFO o logging.ERROR según lo que necesites ver
logging.basicConfig(
    level=LOG_LEVEL,
    format='[%(levelname)s] %(message)s'
)

def print_separator():
    """
    Imprime una línea separadora amarilla en la consola para mejorar la legibilidad.
    """
    print(Fore.YELLOW + '=' * 60)

def print_menu_title(title):
    """
    Imprime un título de menú resaltado con color y separadores.
    """
    print_separator()
    print(Fore.CYAN + Style.BRIGHT + f"*** {title} ***")
    print_separator()

def interactive_folder_selection(start_path=None):
    """
    Permite al usuario navegar por las carpetas de manera interactiva desde la consola.
    Devuelve la ruta seleccionada por el usuario o None si el usuario desea volver al menú principal.
    """
    if start_path is None:
        start_path = os.getcwd()  # Usa la carpeta actual si no se especifica otra
    current_path = os.path.abspath(start_path)
    while True:
        print(Fore.CYAN + f"\n📁 Carpeta actual: {current_path}")
        # Lista solo las subcarpetas de la carpeta actual
        entries = [e for e in os.listdir(current_path) if os.path.isdir(os.path.join(current_path, e))]
        print(Fore.YELLOW + "Subcarpetas disponibles para navegar:")
        for idx, entry in enumerate(entries):
            print(f"  {idx+1}. {entry}")
        print(Fore.MAGENTA + "0. Seleccionar esta carpeta para continuar aquí")
        print(Fore.MAGENTA + "u. Subir un nivel (carpeta anterior)")
        print(Fore.MAGENTA + "m. Volver al menú principal (cancelar selección)")
        print(Fore.MAGENTA + "q. Cerrar la aplicación inmediatamente")
        # Solicita al usuario que seleccione una opción
        sel = input(Fore.WHITE + "Elija una opción (número, 0 para seleccionar, u para subir, m para menú, q para salir): ").strip()
        if sel == '0':
            print(Fore.GREEN + f"✔ Carpeta seleccionada: {current_path}")
            return current_path  # Devuelve la carpeta actual
        elif sel.lower() == 'u':
            parent = os.path.dirname(current_path)
            if parent == current_path:
                print(Fore.RED + "⚠ Ya está en la carpeta raíz, no puede subir más.")
            else:
                current_path = parent  # Sube un nivel en la jerarquía de carpetas
        elif sel.lower() == 'm':
            print(Fore.YELLOW + "↩ Volviendo al menú principal...")
            return None  # Volver al menú principal
        elif sel.lower() == 'q':
            print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
            sys.exit(0)
        elif sel.isdigit():
            idx = int(sel) - 1
            if 0 <= idx < len(entries):
                current_path = os.path.join(current_path, entries[idx])  # Entra en la subcarpeta seleccionada
            else:
                print(Fore.RED + "❌ Índice fuera de rango. Intente con un número válido.")
        else:
            print(Fore.RED + "❌ Opción inválida. Por favor, elija una opción válida.")

def ask_file_paths(prompt_msg, multiple=False, allow_file_dialog=False):
    """
    Solicita al usuario las rutas de los archivos XLSX a procesar.
    Permite seleccionar uno o varios archivos, ya sea por consola o usando un navegador de archivos gráfico.
    Permite volver al menú principal o cerrar la app.
    """
    print(Fore.GREEN + f"\n{prompt_msg}")
    files = []
    if allow_file_dialog:
        print(Fore.YELLOW + "\n¿Cómo desea seleccionar los archivos?")
        print(Fore.MAGENTA + "1. (CLI) Navegar por carpetas desde la consola (modo interactivo)")
        print(Fore.MAGENTA + "2. (GUI) Usar el navegador de archivos gráfico (selección múltiple disponible)")
        print(Fore.MAGENTA + "m. Volver al menú principal (cancelar selección)")
        print(Fore.MAGENTA + "q. Cerrar la aplicación inmediatamente")
        while True:
            sel = input(Fore.WHITE + "Seleccione una opción (1, 2, m, q): ").strip()
            if sel == '1':
                folder = interactive_folder_selection()
                if folder is None:
                    return None  # Volver al menú principal
                files = [os.path.join(folder, f) for f in os.listdir(folder) if f.lower().endswith('.xlsx') and os.path.isfile(os.path.join(folder, f))]
                if not files:
                    print(Fore.RED + "⚠ No se encontraron archivos .xlsx en la carpeta seleccionada.")
                    return []
                print(Fore.GREEN + f"✔ {len(files)} archivo(s) seleccionado(s) para procesar.")
                break
            elif sel == '2':
                root = tk.Tk()
                root.withdraw()  # Oculta la ventana principal de Tkinter
                if multiple:
                    files = filedialog.askopenfilenames(
                        title="Selecciona uno o más archivos XLSX para procesar",
                        filetypes=[("Archivos Excel", "*.xlsx")]
                    )
                    files = list(files)
                else:
                    file = filedialog.askopenfilename(
                        title="Selecciona un archivo XLSX para procesar",
                        filetypes=[("Archivos Excel", "*.xlsx")]
                    )
                    files = [file] if file else []
                root.destroy()
                if not files:
                    print(Fore.RED + "⚠ No se seleccionaron archivos.")
                    return []
                print(Fore.GREEN + f"✔ {len(files)} archivo(s) seleccionado(s) para procesar.")
                break
            elif sel.lower() == 'm':
                print(Fore.YELLOW + "↩ Volviendo al menú principal...")
                return None  # Volver al menú principal
            elif sel.lower() == 'q':
                print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
                sys.exit(0)
            else:
                print(Fore.RED + "❌ Opción inválida. Por favor, elija una opción válida.")
        return files
    # Selección manual por consola
    print(Fore.MAGENTA + "\nIngrese la ruta de archivo o escriba 'm' para menú principal, 'q' para salir.")
    while True:
        path = input(Fore.WHITE + "Ruta de archivo (deje vacío para terminar la selección): ").strip()
        if path.lower() == 'm':
            print(Fore.YELLOW + "↩ Volviendo al menú principal...")
            return None
        if path.lower() == 'q':
            print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
            sys.exit(0)
        if not path:
            if multiple and files:
                print(Fore.GREEN + f"✔ {len(files)} archivo(s) seleccionado(s) para procesar.")
                break
            elif not multiple:
                print(Fore.RED + "⚠ Debe ingresar al menos un archivo.")
                continue
            else:
                print(Fore.RED + "⚠ No se seleccionó ningún archivo.")
                return []
        if not os.path.isfile(path):
            print(Fore.RED + f"❌ Archivo no encontrado: {path}")
            continue
        if not path.lower().endswith('.xlsx'):
            print(Fore.RED + "❌ Solo se permiten archivos con extensión .xlsx")
            continue
        files.append(path)
        if not multiple:
            print(Fore.GREEN + f"✔ Archivo seleccionado: {path}")
            break
    return files

def ask_output_path(default_name):
    """
    Solicita al usuario la carpeta y el nombre del archivo de salida.
    Ofrece tres modos: carpeta actual, navegación CLI, o navegador de archivos GUI.
    Permite volver al menú principal o cerrar la app.
    """
    print(Fore.GREEN + "\n¿Dónde desea guardar el archivo de salida generado por la operación?")
    print(Fore.YELLOW + "Opciones para establecer la carpeta de destino:")
    print(Fore.MAGENTA + "1. (FAST) Guardar en la carpeta actual (deje vacío para usar esta opción)")
    print(Fore.MAGENTA + "2. (CLI) Seleccionar carpeta navegando por consola (modo interactivo)")
    print(Fore.MAGENTA + "3. (GUI) Seleccionar carpeta usando el navegador gráfico")
    print(Fore.MAGENTA + "m. Volver al menú principal (cancelar guardado)")
    print(Fore.MAGENTA + "q. Cerrar la aplicación inmediatamente")
    folder = ""
    while True:
        sel = input(Fore.WHITE + "Seleccione una opción (1, 2, 3, m, q) o deje vacío para carpeta actual: ").strip()
        if not sel or sel == '1':
            folder = os.getcwd()
            print(Fore.GREEN + f"✔ Carpeta actual seleccionada: {folder}")
            break
        elif sel == '2':
            folder = interactive_folder_selection()
            if folder is None:
                print(Fore.YELLOW + "↩ Volviendo al menú principal...")
                return None
            print(Fore.GREEN + f"✔ Carpeta seleccionada: {folder}")
            break
        elif sel == '3':
            root = tk.Tk()
            root.withdraw()
            folder = filedialog.askdirectory(title="Selecciona la carpeta de destino para guardar el archivo")
            root.destroy()
            if not folder:
                print(Fore.RED + "⚠ No se seleccionó ninguna carpeta. Usando carpeta actual.")
                folder = os.getcwd()
            else:
                print(Fore.GREEN + f"✔ Carpeta seleccionada: {folder}")
            break
        elif sel.lower() == 'm':
            print(Fore.YELLOW + "↩ Volviendo al menú principal...")
            return None
        elif sel.lower() == 'q':
            print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
            sys.exit(0)
        else:
            print(Fore.RED + "❌ Opción inválida. Por favor, elija una opción válida.")
    if not os.path.isdir(folder):
        print(Fore.RED + "⚠ La carpeta no existe. Usando carpeta actual.")
        folder = os.getcwd()
    name = input(Fore.WHITE + f"Ingrese el nombre del archivo de salida (sin extensión, por defecto '{default_name}'): ").strip()
    if name.lower() == 'm':
        print(Fore.YELLOW + "↩ Volviendo al menú principal...")
        return None
    if name.lower() == 'q':
        print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
        sys.exit(0)
    if not name:
        name = default_name
    print(Fore.GREEN + f"✔ El archivo se guardará como: {name}.xlsx")
    return os.path.join(folder, name + '.xlsx')

def get_headers(ws):
    """
    Obtiene los encabezados (primera fila) de una hoja de Excel.
    """
    return [cell.value for cell in ws[1]]

def check_same_headers(file_paths):
    """
    Verifica que todos los archivos tengan los mismos encabezados.
    Si encuentra diferencias, muestra un error y retorna False.
    """
    headers = None
    for path in file_paths:
        wb = load_workbook(path, read_only=True)
        ws = wb.active
        current_headers = get_headers(ws)
        if headers is None:
            headers = current_headers
        elif headers != current_headers:
            logging.error(f"Encabezados diferentes en el archivo: {path}")
            return False
        wb.close()
    return True

# --- SECCIÓN DE REPORTES DETALLADOS ---
def print_report(report_data):
    """
    Imprime un reporte detallado de la operación realizada, incluyendo estadísticas y uso de recursos.
    """
    print_separator()
    print(Fore.GREEN + Style.BRIGHT + "📊 REPORTE DETALLADO DE LA OPERACIÓN")
    print_separator()
    print(Fore.CYAN + f"Archivos procesados: {report_data.get('files_processed', '-')}")
    # Mostrar líneas por archivo si está disponible
    file_lines = report_data.get('file_lines', None)
    if file_lines:
        print(Fore.BLUE + "Líneas procesadas por archivo:")
        for fname, lines in file_lines.items():
            print(Fore.BLUE + f"  - {os.path.basename(fname)}: {lines} líneas")
    # Mostrar líneas al inicio y al final con textos personalizados
    print(Fore.CYAN + f"Líneas al inicio: {report_data.get('lines_in_text', '-')}")
    print(Fore.CYAN + f"Líneas al final: {report_data.get('lines_out_text', '-')}")
    print(Fore.CYAN + f"Tiempo total de operación: {report_data.get('duration', '-')} segundos")
    print(Fore.CYAN + f"Tamaño del archivo de salida: {report_data.get('output_size_kb', '-')} KB")
    print(Fore.CYAN + f"Ruta del archivo de salida: {report_data.get('output_path', '-')}")
    print(Fore.CYAN + f"Carpeta de destino: {report_data.get('output_folder', '-')}")
    print(Fore.CYAN + f"Memoria RAM utilizada: {report_data.get('ram_used_mb', '-')} MB")
    print(Fore.CYAN + f"Porcentaje de CPU utilizado: {report_data.get('cpu_percent', '-')} %")
    print_separator()
    print(Fore.YELLOW + "✅ Operación finalizada. Revise el archivo generado y los detalles anteriores para más información.")

# --- FIN SECCIÓN DE REPORTES DETALLADOS ---

def merge_xlsx():
    """
    Función principal para combinar (merge) varios archivos XLSX en uno solo.
    - Solicita al usuario los archivos a combinar.
    - Verifica que tengan la misma estructura.
    - Copia los datos de todos los archivos en uno nuevo.
    - Mide recursos y tiempo, y muestra un reporte detallado.
    """
    print_menu_title("Función MERGE - Consolidar archivos XLSX")
    files = ask_file_paths(
        "Selecciona los archivos XLSX a combinar (varios archivos con misma estructura):",
        multiple=True,
        allow_file_dialog=True
    )
    if files is None:
        return  # Volver al menú principal
    if not files:
        print(Fore.RED + "⚠ No se seleccionaron archivos para combinar. Operación cancelada.")
        return
    if not check_same_headers(files):
        print(Fore.RED + "❌ Los archivos seleccionados no tienen los mismos encabezados. Operación cancelada.")
        return
    output_path = ask_output_path("merge_result")
    if output_path is None:
        return  # Volver al menú principal
    # --- INICIO MEDICIÓN DE RECURSOS Y TIEMPO ---
    start_time = time.time()
    process = psutil.Process(os.getpid())
    ram_before = process.memory_info().rss
    cpu_before = psutil.cpu_percent(interval=None)
    # ---
    try:
        wb_out = Workbook()  # Crea un nuevo libro de Excel para el resultado
        ws_out = wb_out.active
        # Copia el encabezado del primer archivo
        wb_first = load_workbook(files[0], read_only=True)
        ws_first = wb_first.active
        ws_out.append(get_headers(ws_first))
        wb_first.close()
        # Copia los datos de todos los archivos (ignorando la primera fila)
        total_lines_in = 0
        file_lines = {}
        for file in files:
            logging.debug(f"Procesando archivo: {file}")
            wb = load_workbook(file, read_only=True)
            ws = wb.active
            file_line_count = sum(1 for _ in ws.iter_rows(min_row=2, values_only=True))
            file_lines[file] = file_line_count
            total_lines_in += file_line_count
            # Volver a iterar para copiar las filas (no se puede reutilizar el iterador anterior)
            for row in ws.iter_rows(min_row=2, values_only=True):
                ws_out.append(row)
            wb.close()
        wb_out.save(output_path)  # Guarda el archivo combinado
        # --- FIN MEDICIÓN DE RECURSOS Y TIEMPO ---
        end_time = time.time()
        ram_after = process.memory_info().rss
        cpu_after = psutil.cpu_percent(interval=None)
        output_size_kb = os.path.getsize(output_path) // 1024
        output_folder = os.path.dirname(output_path)
        # Cuenta las filas finales (sin encabezado)
        wb_check = load_workbook(output_path, read_only=True)
        ws_check = wb_check.active
        lines_out = ws_check.max_row - 1  # sin encabezado
        wb_check.close()
        # Prepara los datos para el reporte detallado
        report_data = {
            'files_processed': len(files),
            'file_lines': file_lines,
            'lines_in': total_lines_in,
            'lines_in_text': f"{total_lines_in} Líneas Totales (antes de merge)",
            'lines_out': lines_out,
            'lines_out_text': f"{lines_out} Líneas combinadas (sin duplicados de encabezado)",
            'duration': round(end_time - start_time, 2),
            'output_size_kb': output_size_kb,
            'output_path': output_path,
            'output_folder': output_folder,
            'ram_used_mb': round((ram_after - ram_before) / (1024*1024), 2),
            'cpu_percent': cpu_after
        }
        print(Fore.GREEN + f"\n🎉 Archivos combinados exitosamente en: {output_path}")
        print_report(report_data)
    except Exception as e:
        logging.error(f"❌ Error al combinar archivos: {e}")

def wipe_xlsx():
    """
    Función principal para eliminar duplicados de un archivo XLSX.
    - Solicita al usuario el archivo a purgar.
    - Pide la columna que contiene los valores únicos.
    - Elimina filas duplicadas según esa columna.
    - Mide recursos y tiempo, y muestra un reporte detallado.
    """
    print_menu_title("Función WIPE - Eliminar duplicados en archivo XLSX")
    files = ask_file_paths(
        "Selecciona el archivo XLSX a purgar (archivo unico):",
        multiple=False,
        allow_file_dialog=True
    )
    if files is None:
        return  # Volver al menú principal
    if not files:
        print(Fore.RED + "⚠ No se seleccionó archivo para purgar. Operación cancelada.")
        return
    file = files[0]
    # --- INICIO MEDICIÓN DE RECURSOS Y TIEMPO ---
    start_time = time.time()
    process = psutil.Process(os.getpid())
    ram_before = process.memory_info().rss
    cpu_before = psutil.cpu_percent(interval=None)
    # ---
    try:
        wb = load_workbook(file)
        ws = wb.active
        headers = get_headers(ws)
        # Muestra las columnas disponibles y sus encabezados
        print(Fore.BLUE + "\nColumnas disponibles en el archivo:")
        for idx, header in enumerate(headers):
            col_letter = get_column_letter(idx + 1)
            print(f"  Columna {col_letter} - {header}")
        # Solicita al usuario la columna que contiene los valores únicos
        while True:
            col_input = input(Fore.WHITE + "¿Qué columna contiene los valores únicos? (Letra de columna, 'm' menú, 'q' salir): ").strip().upper()
            if col_input.lower() == 'm':
                print(Fore.YELLOW + "↩ Volviendo al menú principal...")
                return  # Volver al menú principal
            if col_input.lower() == 'q':
                print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
                sys.exit(0)
            if not col_input:
                print(Fore.RED + "⚠ Debe ingresar una letra de columna (ejemplo: A, B, C...).")
                continue
            try:
                col_idx = ord(col_input) - ord('A')
                if 0 <= col_idx < len(headers):
                    print(Fore.GREEN + f"✔ Columna seleccionada: {col_input} - {headers[col_idx]}")
                    break
                else:
                    print(Fore.RED + "❌ Columna fuera de rango. Intente con una letra válida.")
            except Exception:
                print(Fore.RED + "❌ Entrada inválida. Intente nuevamente.")
        unique_col = col_idx
        seen = set()  # Conjunto para almacenar los valores únicos encontrados
        rows_to_keep = []
        total_lines_in = ws.max_row - 1  # sin encabezado
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                rows_to_keep.append(row)  # Siempre guarda el encabezado
                continue
            val = row[unique_col]
            if val not in seen:
                seen.add(val)
                rows_to_keep.append(row)
        output_path = ask_output_path("wipe_result")
        if output_path is None:
            return  # Volver al menú principal
        wb_out = Workbook()
        ws_out = wb_out.active
        for row in rows_to_keep:
            ws_out.append(row)
        wb_out.save(output_path)
        # --- FIN MEDICIÓN DE RECURSOS Y TIEMPO ---
        end_time = time.time()
        ram_after = process.memory_info().rss
        cpu_after = psutil.cpu_percent(interval=None)
        output_size_kb = os.path.getsize(output_path) // 1024
        output_folder = os.path.dirname(output_path)
        lines_out = len(rows_to_keep) - 1  # sin encabezado
        lines_removed = total_lines_in - lines_out
        # Prepara los datos para el reporte detallado
        report_data = {
            'files_processed': 1,
            'file_lines': {file: total_lines_in},
            'lines_in': total_lines_in,
            'lines_in_text': f"{total_lines_in} Líneas Totales (antes de purgar)",
            'lines_out': lines_out,
            'lines_out_text': f"{lines_out} Líneas después del Wipe - {lines_removed} duplicados eliminados",
            'duration': round(end_time - start_time, 2),
            'output_size_kb': output_size_kb,
            'output_path': output_path,
            'output_folder': output_folder,
            'ram_used_mb': round((ram_after - ram_before) / (1024*1024), 2),
            'cpu_percent': cpu_after
        }
        print(Fore.GREEN + f"\n🧹 Archivo purgado exitosamente en: {output_path}")
        print_report(report_data)
    except Exception as e:
        logging.error(f"❌ Error al purgar archivo: {e}")

def main_menu():
    """
    Muestra el menú principal del programa y gestiona la selección del usuario.
    Permite elegir entre combinar archivos, eliminar duplicados o salir.
    """
    while True:
        print_menu_title("MERGE-WIPER - Menú Principal")
        print(Fore.MAGENTA + "1. Merge (Combinar/Mergear archivos XLSX) - Une varios archivos Excel en uno solo.")
        print(Fore.MAGENTA + "2. Wipe (Eliminar/Wipear duplicados en archivo XLSX) - Elimina filas duplicadas según una columna.")
        print(Fore.MAGENTA + "3. Salir del programa")
        print(Fore.YELLOW + "Puede presionar 'q' en cualquier momento para cerrar la aplicación.")
        choice = input(Fore.WHITE + "Seleccione una opción (1, 2, 3 o 'q'): ").strip()
        if choice == '1':
            merge_xlsx()
        elif choice == '2':
            wipe_xlsx()
        elif choice == '3':
            print(Fore.YELLOW + "👋 ¡Hasta luego! Gracias por usar Merge-Wiper.")
            break
        elif choice.lower() == 'q':
            print(Fore.YELLOW + "⏹ Cerrando la aplicación. ¡Hasta pronto!")
            sys.exit(0)
        else:
            print(Fore.RED + "❌ Opción inválida. Por favor, seleccione una opción válida.")

if __name__ == "__main__":
    # Punto de entrada principal del programa
    try:
        main_menu()
    except KeyboardInterrupt:
        print(Fore.YELLOW + "\n⏹ Operación cancelada por el usuario. Puede reiniciar el programa cuando lo desee.")
